from datetime import datetime
import os
import numpy as np
import math
from flask import (
    Flask,
    render_template,
    request,
    flash,
    redirect,
    url_for,
    session,
    Response,
    send_file,
)
import uuid
import traceback
import io 

import matplotlib
matplotlib.use("Agg") 
import matplotlib.pyplot as plt
import seaborn as sns

from openpyxl import Workbook, load_workbook 
from openpyxl.styles import Font, Alignment
from openpyxl.cell.cell import MergedCell

# --- REPORTLAB IMPORTS ---
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Image,
    Table,
    TableStyle,
    PageBreak,
    KeepInFrame,
    HRFlowable,
)
from reportlab.platypus.flowables import KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.lib.units import inch, cm
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import A4
from reportlab.lib.fonts import addMapping

# --- END REPORTLAB IMPORTS ---
try:
    from reportlab.lib.units import pt
except ImportError:
    print("DEBUG ReportLab: 'pt' not found directly in reportlab.lib.units. Defining manually.")
    pt = inch / 72.0
# --- KẾT THÚC TỰ ĐỊNH NGHĨA 'pt' ---

from model.model import (
    get_criteria_from_db,
    get_all_alternatives,
    add_criteria,
    add_alternative,
    save_criteria_weights,
    save_alternative_scores,
    save_criteria_comparison_matrix,
    get_criteria_by_ids,
    get_alternatives_by_ids,
    get_or_create_session_db_id,
    save_ahp_analysis,
    get_ahp_analyses_by_session_db_id,
    get_ahp_analysis_by_id,
)
from controller.ahp import (
    get_sorted_criteria_with_weights,
    parse_saaty_value,
    calculate_ahp,
)

app = Flask(__name__)
app.secret_key = b'_5#y2L"F4Q8z\n\xec]/'
app.jinja_env.globals.update(enumerate=enumerate)

# --- ĐĂNG KÝ FONT VỚI PDFMETRICS (QUAN TRỌNG CHO REPORTLAB) ---
FONT_NAME_REGULAR = "MyDejaVuSans"
FONT_NAME_BOLD = "MyDejaVuSansBold"

try:
    static_dir = os.path.join(app.root_path, "static")
    font_path_regular_actual = os.path.join(static_dir, "fonts", "ttf", "DejaVuSans.ttf")
    font_path_bold_actual = os.path.join(static_dir, "fonts", "ttf", "DejaVuSans-Bold.ttf")

    font_registered_regular = False
    font_registered_bold = False

    if os.path.exists(font_path_regular_actual):
        pdfmetrics.registerFont(TTFont(FONT_NAME_REGULAR, font_path_regular_actual))
        font_registered_regular = True
    else:
        print(f"ERROR ReportLab: Font file for '{FONT_NAME_REGULAR}' not found at {font_path_regular_actual}.")

    if os.path.exists(font_path_bold_actual):
        pdfmetrics.registerFont(TTFont(FONT_NAME_BOLD, font_path_bold_actual))
        font_registered_bold = True
    else:
        print(f"ERROR ReportLab: Font file for '{FONT_NAME_BOLD}' not found at {font_path_bold_actual}.")

    if font_registered_regular:
        addMapping(FONT_NAME_REGULAR, 0, 0, FONT_NAME_REGULAR)
        if font_registered_bold:
            addMapping(FONT_NAME_REGULAR, 1, 0, FONT_NAME_BOLD)
        else:
            addMapping(FONT_NAME_REGULAR, 1, 0, FONT_NAME_REGULAR)
    else:
        print("ERROR ReportLab: Regular font not registered, skipping mappings.")

except Exception as e:
    print(f"ERROR ReportLab: Failed to register fonts or add mappings: {e}")
    traceback.print_exc()


# --- CÁC HẰNG SỐ CHO IMPORT EXCEL ---
SHEET_NAME_LISTS = "DanhSach"
SHEET_NAME_CRITERIA_MATRIX = "MaTranTieuChi"
SHEET_NAME_ALL_ALT_MATRICES = "TatCaMaTranPA"
CRITERION_BLOCK_MARKER = "Tiêu chí: "
ALLOWED_EXTENSIONS = {'xlsx'}

CHARTS_FOLDER = os.path.join(app.static_folder, "generated_charts")
if not os.path.exists(CHARTS_FOLDER):
    os.makedirs(CHARTS_FOLDER)


# --- CÁC HÀM HELPER ---
def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def parse_matrix_from_sheet(sheet, expected_headers, expected_row_labels, sheet_title_for_error=""):
    """
    Hàm chung để đọc ma trận từ một sheet (cho MaTranTieuChi).
    Trả về ma trận dạng numpy array hoặc None nếu có lỗi.
    """
    if not sheet_title_for_error:
        sheet_title_for_error = sheet.title

    headers_excel = [cell.value for cell in sheet[1][1:] if cell.value is not None] # Bỏ ô A1, chỉ lấy cell có giá trị
    row_labels_excel = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1) if sheet.cell(row=i, column=1).value is not None]


    if not headers_excel or not row_labels_excel:
        flash(f"Sheet '{sheet_title_for_error}' thiếu header hoặc row labels.", "error")
        return None

    # Kiểm tra xem tất cả expected_headers có trong headers_excel không và ngược lại
    if sorted(headers_excel) != sorted(expected_headers):
        missing_in_excel = set(expected_headers) - set(headers_excel)
        extra_in_excel = set(headers_excel) - set(expected_headers)
        msg = f"Headers trong sheet '{sheet_title_for_error}' không khớp. "
        if missing_in_excel: msg += f"Thiếu: {missing_in_excel}. "
        if extra_in_excel: msg += f"Thừa: {extra_in_excel}."
        flash(msg, "error")
        return None

    if sorted(row_labels_excel) != sorted(expected_row_labels):
        missing_in_excel = set(expected_row_labels) - set(row_labels_excel)
        extra_in_excel = set(row_labels_excel) - set(expected_row_labels)
        msg = f"Row labels trong sheet '{sheet_title_for_error}' không khớp. "
        if missing_in_excel: msg += f"Thiếu: {missing_in_excel}. "
        if extra_in_excel: msg += f"Thừa: {extra_in_excel}."
        flash(msg, "error")
        return None

    header_map_excel = {name: i for i, name in enumerate(headers_excel)}
    row_label_map_excel = {name: i for i, name in enumerate(row_labels_excel)}

    n = len(expected_headers)
    matrix_str = np.ones((n, n), dtype=object)

    for r_idx_expected, r_name_expected in enumerate(expected_headers): # Dùng expected_headers cho cả row và col
        excel_r_idx = row_label_map_excel.get(r_name_expected)
        if excel_r_idx is None: # Lỗi logic, đã check ở trên
             flash(f"Lỗi nội bộ: Không tìm thấy row label '{r_name_expected}' trong map Excel.", "error"); return None

        for c_idx_expected, c_name_expected in enumerate(expected_headers):
            excel_c_idx = header_map_excel.get(c_name_expected)
            if excel_c_idx is None: # Lỗi logic
                flash(f"Lỗi nội bộ: Không tìm thấy header '{c_name_expected}' trong map Excel.", "error"); return None
            
            cell_value = sheet.cell(row=excel_r_idx + 2, column=excel_c_idx + 2).value
            matrix_str[r_idx_expected, c_idx_expected] = str(cell_value) if cell_value is not None else "1"

    parsed_matrix_values = np.ones((n, n), dtype=float)
    try:
        for i in range(n):
            for j in range(n):
                value_s = matrix_str[i, j]
                if i == j:
                    parsed_matrix_values[i, j] = parse_saaty_value(value_s.strip(), is_diagonal=True)
                elif i < j:
                    parsed_val = parse_saaty_value(value_s.strip())
                    parsed_matrix_values[i, j] = parsed_val
                    parsed_matrix_values[j, i] = 1.0 / parsed_val if abs(parsed_val) > 1e-9 else np.inf
    except ValueError as e:
        flash(f"Lỗi giá trị Saaty không hợp lệ trong sheet '{sheet_title_for_error}': {e}", "error")
        return None
    return parsed_matrix_values

def parse_single_matrix_block(sheet, start_row_header, num_alternatives, alternative_names_ordered, sheet_title_for_error=""):
    """
    Đọc một khối ma trận phương án từ sheet `TatCaMaTranPA`.
    """
    if not sheet_title_for_error:
        sheet_title_for_error = sheet.title

    # Header của ma trận Phương án (tên Phương án) nằm ở các cột B, C,... của dòng start_row_header
    headers_excel = [sheet.cell(row=start_row_header, column=j + 2).value for j in range(num_alternatives) if sheet.cell(row=start_row_header, column=j + 2).value is not None]
    # Row labels của ma trận Phương án (tên Phương án) nằm ở cột A, từ dòng start_row_header + 1
    row_labels_excel = [sheet.cell(row=start_row_header + 1 + i, column=1).value for i in range(num_alternatives) if sheet.cell(row=start_row_header + 1 + i, column=1).value is not None]

    if len(headers_excel) != num_alternatives:
        flash(f"Số lượng header Phương án ({len(headers_excel)}) tại dòng {start_row_header} sheet '{sheet_title_for_error}' không khớp với số lượng phương án ({num_alternatives}). Headers: {headers_excel}", "error")
        return None
    if len(row_labels_excel) != num_alternatives:
        flash(f"Số lượng row label Phương án ({len(row_labels_excel)}) từ dòng {start_row_header+1} sheet '{sheet_title_for_error}' không khớp với số lượng phương án ({num_alternatives}). Labels: {row_labels_excel}", "error")
        return None

    if sorted(headers_excel) != sorted(alternative_names_ordered):
        flash(f"Headers Phương án tại dòng {start_row_header} sheet '{sheet_title_for_error}' ({headers_excel}) không khớp với danh sách phương án ({alternative_names_ordered}).", "error")
        return None
    if sorted(row_labels_excel) != sorted(alternative_names_ordered):
        flash(f"Row labels Phương án từ dòng {start_row_header+1} sheet '{sheet_title_for_error}' ({row_labels_excel}) không khớp với danh sách phương án ({alternative_names_ordered}).", "error")
        return None

    header_map_excel = {name: i for i, name in enumerate(headers_excel)}
    row_label_map_excel = {name: i for i, name in enumerate(row_labels_excel)}

    matrix_str = np.ones((num_alternatives, num_alternatives), dtype=object)
    for r_idx_expected, r_name_expected in enumerate(alternative_names_ordered):
        excel_r_idx = row_label_map_excel.get(r_name_expected)
        if excel_r_idx is None: flash(f"Lỗi nội bộ: Không tìm thấy Phương án '{r_name_expected}' trong row labels map.", "error"); return None

        for c_idx_expected, c_name_expected in enumerate(alternative_names_ordered):
            excel_c_idx = header_map_excel.get(c_name_expected)
            if excel_c_idx is None: flash(f"Lỗi nội bộ: Không tìm thấy Phương án '{c_name_expected}' trong headers map.", "error"); return None
            
            cell_value = sheet.cell(row=start_row_header + 1 + excel_r_idx, column=2 + excel_c_idx).value
            matrix_str[r_idx_expected, c_idx_expected] = str(cell_value) if cell_value is not None else "1"
            
    parsed_matrix_values = np.ones((num_alternatives, num_alternatives), dtype=float)
    try:
        for i in range(num_alternatives):
            for j in range(num_alternatives):
                value_s = matrix_str[i,j]
                if i == j:
                    parsed_matrix_values[i, j] = parse_saaty_value(value_s.strip(), is_diagonal=True)
                elif i < j:
                    parsed_val = parse_saaty_value(value_s.strip())
                    parsed_matrix_values[i, j] = parsed_val
                    parsed_matrix_values[j, i] = 1.0 / parsed_val if abs(parsed_val) > 1e-9 else np.inf
    except ValueError as e:
        flash(f"Lỗi giá trị Saaty không hợp lệ trong ma trận Phương án (bắt đầu dòng header {start_row_header}) sheet '{sheet_title_for_error}': {e}", "error")
        return None
    return parsed_matrix_values

def convert_numpy_matrix_to_form_data(matrix_np, form_prefix, crit_idx=None):
    form_data = {}
    n_rows, n_cols = matrix_np.shape
    for r in range(n_rows):
        for c in range(n_cols):
            key_base = f"[{r}][{c}]"
            key = f"{form_prefix}{key_base}" if crit_idx is None else f"{form_prefix}[{crit_idx}]{key_base}"
            
            if r == c:
                form_data[key] = "1"
            elif r < c: # Chỉ điền nửa trên cho form
                value = matrix_np[r, c]
                if 0 < value < 1:
                    inv_value = 1.0 / value
                    if abs(inv_value - round(inv_value)) < 1e-6 :
                         form_data[key] = f"1/{int(round(inv_value))}"
                    else:
                         form_data[key] = f"{value:.4f}".rstrip('0').rstrip('.')
                elif value >=1:
                    form_data[key] = str(int(value)) if abs(value - round(value)) < 1e-6 else f"{value:.4f}".rstrip('0').rstrip('.')
                else: # value = 0 or inf, ...
                    form_data[key] = str(value) # Để thể hiện lỗi nếu có
    return form_data
def convert_all_alt_matrices_to_form_data(imported_alternative_matrices, criteria_tuples, alternatives_tuples):
    """
    Chuyển đổi tất cả các ma trận phương án (import từ Excel) thành dict form_data cho render template.
    imported_alternative_matrices: dict {str(crit_idx): matrix_list}
    criteria_tuples: [(id, name), ...]
    alternatives_tuples: [(id, name), ...]
    """
    form_data = {}
    n_criteria = len(criteria_tuples)
    n_alternatives = len(alternatives_tuples)
    for crit_idx_str, matrix_list in imported_alternative_matrices.items():
        crit_idx = int(crit_idx_str)
        matrix_np = np.array(matrix_list)
        # Sử dụng hàm convert_numpy_matrix_to_form_data đã có
        form_data.update(convert_numpy_matrix_to_form_data(matrix_np, "alt_matrix", crit_idx))
    return form_data
# --- CÁC ROUTE ---
@app.route("/", methods=["GET"])
def home():
    if "flask_session_id" not in session:
        session["flask_session_id"] = str(uuid.uuid4())
    keys_to_pop = [
        "criteria_ahp_results", "current_criteria_tuples", "current_alternatives_tuples",
        "form_data_criteria_temp", "form_data_alternatives_temp", "alternative_crs_temp",
        "data_imported_from_excel", "imported_criteria_matrix", "imported_alternative_matrices"
    ]
    for key in keys_to_pop:
        session.pop(key, None)
    return render_template(
        "matrix.html",
        criteria=get_criteria_from_db(), # Dùng cho form chọn thủ công ban đầu
        alternatives=get_all_alternatives(), # Dùng cho form chọn thủ công
        show_input_forms=True,
        show_ahp_steps=False,
        show_alternatives_form=False,
    )

# ... add_criteria_route, add_alternative_route, start_ahp_route (GIỮ NGUYÊN) ...
@app.route("/add_criteria", methods=["POST"])
def add_criteria_route():
    crit_name = request.form.get("criteria_name", "").strip()
    if not crit_name:
        flash("Tên tiêu chí không được để trống.", "error")
    else:
        if add_criteria(crit_name): # Giả sử hàm này trả về True nếu thành công
            flash(f"Đã thêm tiêu chí '{crit_name}'.", "success")
        else:
            flash(f"Tiêu chí '{crit_name}' có thể đã tồn tại hoặc có lỗi khi thêm.", "error")
    return redirect(url_for("home"))


@app.route("/add_alternative", methods=["POST"])
def add_alternative_route():
    alt_name = request.form.get("alternative_name", "").strip()
    if not alt_name:
        flash("Tên phương án không được để trống.", "error")
    else:
        if add_alternative(alt_name): # Giả sử hàm này trả về True nếu thành công
            flash(f"Đã thêm phương án '{alt_name}'.", "success")
        else:
            flash(f"Phương án '{alt_name}' có thể đã tồn tại hoặc có lỗi khi thêm.", "error")
    return redirect(url_for("home"))

@app.route("/start_ahp", methods=["POST"]) # Khi người dùng chọn TC, PA thủ công và bấm "Bắt đầu AHP"
def start_ahp_route():
    if "flask_session_id" not in session:
        flash("Lỗi phiên làm việc. Vui lòng thử lại.", "error")
        return redirect(url_for("home"))
    try:
        selected_criteria_ids = [int(cid) for cid in request.form.getlist("selected_criteria_ids")]
        selected_alternative_ids = [int(aid) for aid in request.form.getlist("selected_alternative_ids")]
    except ValueError:
        flash("ID tiêu chí hoặc phương án không hợp lệ.", "error")
        return redirect(url_for("home"))

    MIN_CRITERIA = 3
    MIN_ALTERNATIVES = 2
    if len(selected_criteria_ids) < MIN_CRITERIA:
        flash(f"Cần chọn ít nhất {MIN_CRITERIA} tiêu chí.", "error")
        return redirect(url_for("home"))
    if len(selected_alternative_ids) < MIN_ALTERNATIVES:
        flash(f"Cần chọn ít nhất {MIN_ALTERNATIVES} phương án.", "error")
        return redirect(url_for("home"))

    # Lấy thông tin chi tiết từ DB dựa trên ID đã chọn
    # Giả định các hàm get_criteria_by_ids và get_alternatives_by_ids trả về list of tuples [(id, name), ...]
    selected_criteria_tuples = get_criteria_by_ids(selected_criteria_ids)
    selected_alternative_tuples = get_alternatives_by_ids(selected_alternative_ids)

    if len(selected_criteria_tuples) != len(selected_criteria_ids) or \
       len(selected_alternative_tuples) != len(selected_alternative_ids):
        flash("Lỗi khi lấy dữ liệu tiêu chí/phương án từ cơ sở dữ liệu.", "error")
        return redirect(url_for("home"))

    # Lưu vào session
    session['current_criteria_tuples'] = selected_criteria_tuples
    session['current_alternatives_tuples'] = selected_alternative_tuples
    
    # Xóa các dữ liệu AHP cũ nếu có
    keys_to_pop = ["criteria_ahp_results", "form_data_criteria_temp", "form_data_alternatives_temp", "alternative_crs_temp",
                   "data_imported_from_excel", "imported_criteria_matrix", "imported_alternative_matrices"]
    for key in keys_to_pop:
        session.pop(key, None)

    flash(f"Bắt đầu AHP với {len(selected_criteria_tuples)} tiêu chí và {len(selected_alternative_tuples)} phương án.", "info")
    return render_template(
        "matrix.html",
        criteria=selected_criteria_tuples,
        alternatives=selected_alternative_tuples,
        show_input_forms=False, # Ẩn form chọn ban đầu
        show_ahp_steps=True,    # Hiển thị các bước AHP
        show_alternatives_form=False, # Bước 1 (ma trận TC) trước
        imported_data=False # Không phải từ import
    )

@app.route('/import_excel', methods=['POST'])
def import_excel_route():
    if 'excel_file' not in request.files:
        flash('Không có file Excel nào được chọn.', 'error')
        return redirect(url_for('home'))
    file = request.files['excel_file']
    if file.filename == '':
        flash('Không có file Excel nào được chọn.', 'error')
        return redirect(url_for('home'))

    if file and allowed_file(file.filename):
        try:
            workbook = load_workbook(filename=io.BytesIO(file.read()))

            if SHEET_NAME_LISTS not in workbook.sheetnames:
                flash(f"Không tìm thấy sheet '{SHEET_NAME_LISTS}' trong file Excel.", "error"); return redirect(url_for('home'))
            
            sheet_lists = workbook[SHEET_NAME_LISTS]
            criteria_names = [row[0].value for row in sheet_lists.iter_rows(min_row=2, max_col=1) if row[0].value]
            alternative_names = [row[0].value for row in sheet_lists.iter_rows(min_row=2, min_col=2, max_col=2) if row[0].value]
            # Lấy ID thực tế từ DB
            criteria_db = get_criteria_from_db()  # [(id, name), ...]
            criteria_name_to_id = {name: id for id, name in criteria_db}
            alternatives_db = get_all_alternatives()
            alternative_name_to_id = {name: id for id, name in alternatives_db}

            if not criteria_names or len(criteria_names) < 4:
                flash('Cần ít nhất 4 tiêu chí trong sheet "DanhSach".', 'error'); return redirect(url_for('home'))
            if not alternative_names or len(alternative_names) < 3:
                flash('Cần ít nhất 3 phương án trong sheet "DanhSach".', 'error'); return redirect(url_for('home'))
            
            # Dùng đúng ID thực tế
            session['current_criteria_tuples'] = [(criteria_name_to_id[name], name) for name in criteria_names if name in criteria_name_to_id]
            session['current_alternatives_tuples'] = [(alternative_name_to_id[name], name) for name in alternative_names if name in alternative_name_to_id]
            num_alternatives = len(alternative_names)

            if SHEET_NAME_CRITERIA_MATRIX not in workbook.sheetnames:
                flash(f"Không tìm thấy sheet '{SHEET_NAME_CRITERIA_MATRIX}'.", "error"); return redirect(url_for('home'))
            sheet_crit_matrix = workbook[SHEET_NAME_CRITERIA_MATRIX]
            criteria_matrix_np = parse_matrix_from_sheet(sheet_crit_matrix, criteria_names, criteria_names)
            if criteria_matrix_np is None: return redirect(url_for('home'))
            session['imported_criteria_matrix'] = criteria_matrix_np.tolist()

            if SHEET_NAME_ALL_ALT_MATRICES not in workbook.sheetnames:
                flash(f"Không tìm thấy sheet '{SHEET_NAME_ALL_ALT_MATRICES}'.", "error"); return redirect(url_for('home'))
            sheet_all_pa = workbook[SHEET_NAME_ALL_ALT_MATRICES]
            imported_alternative_matrices = {}
            found_criteria_in_pa_sheet = set()
            current_row = 1
            while current_row <= sheet_all_pa.max_row:
                cell_value = sheet_all_pa.cell(row=current_row, column=1).value
                if cell_value and isinstance(cell_value, str) and cell_value.startswith(CRITERION_BLOCK_MARKER):
                    criterion_name_from_marker = cell_value[len(CRITERION_BLOCK_MARKER):].strip()
                    if criterion_name_from_marker not in criteria_names:
                        flash(f"Tên tiêu chí '{criterion_name_from_marker}' (dòng {current_row}) trong '{SHEET_NAME_ALL_ALT_MATRICES}' không khớp '{SHEET_NAME_LISTS}'.", "error")
                        return redirect(url_for('home'))
                    crit_idx = criteria_names.index(criterion_name_from_marker)
                    found_criteria_in_pa_sheet.add(criterion_name_from_marker)
                    
                    alt_matrix_np = parse_single_matrix_block(sheet_all_pa, current_row + 1, num_alternatives, alternative_names)
                    if alt_matrix_np is None: return redirect(url_for('home'))
                    imported_alternative_matrices[str(crit_idx)] = alt_matrix_np.tolist()
                    current_row += (1 + num_alternatives) 
                current_row += 1
            
            if len(found_criteria_in_pa_sheet) != len(criteria_names):
                missing = set(criteria_names) - found_criteria_in_pa_sheet
                flash(f"Thiếu ma trận Phương án cho các tiêu chí: {', '.join(missing)} trong sheet '{SHEET_NAME_ALL_ALT_MATRICES}'.", "error")
                return redirect(url_for('home'))

            session['imported_alternative_matrices'] = imported_alternative_matrices
            session['data_imported_from_excel'] = True
            flash('Đã nhập dữ liệu từ Excel thành công!', 'success')
            return render_template(
                "matrix.html",
                criteria=session['current_criteria_tuples'],
                alternatives=session['current_alternatives_tuples'],
                show_input_forms=False, show_ahp_steps=True, show_alternatives_form=False,
                form_data_criteria=convert_numpy_matrix_to_form_data(criteria_matrix_np, "matrix"),
                imported_data=True
            )
        except Exception as e:
            flash(f'Lỗi khi xử lý file Excel: {e}', 'error')
            traceback.print_exc()
            keys_to_clean = ["data_imported_from_excel", "imported_criteria_matrix", "imported_alternative_matrices",
                             "current_criteria_tuples", "current_alternatives_tuples"]
            for key in keys_to_clean: session.pop(key, None)
            return redirect(url_for('home'))
    else:
        flash('Loại file không hợp lệ. Chỉ chấp nhận file .xlsx.', 'error')
        return redirect(url_for('home'))


@app.route("/calculate_criteria", methods=["POST"])
def calculate_criteria_route():
    flask_session_id_value = session.get("flask_session_id")
    current_session_db_id = get_or_create_session_db_id(flask_session_id_value)
    db_criteria_tuples = session.get("current_criteria_tuples")
    db_alternatives_tuples = session.get("current_alternatives_tuples")

    if not all([current_session_db_id, db_criteria_tuples, db_alternatives_tuples]):
        flash("Lỗi phiên làm việc hoặc thiếu dữ liệu. Vui lòng bắt đầu lại.", "error")
        return redirect(url_for("home"))

    n_criteria = len(db_criteria_tuples)
    criteria_names_ordered = [name for _, name in db_criteria_tuples]
    form_data_criteria_to_render = {}
    criteria_matrix_parsed_np = None

    # Chuẩn bị form_data_alternatives trước, bất kể kết quả của ma trận TC
    # để nếu TC nhất quán, dữ liệu này đã sẵn sàng để truyền đi
    form_data_alt_to_render = {}
    if session.get('data_imported_from_excel') and session.get('imported_alternative_matrices'):
        form_data_alt_to_render = convert_all_alt_matrices_to_form_data(
            session.get('imported_alternative_matrices'),
            db_criteria_tuples,
            db_alternatives_tuples
        )
    elif session.get("form_data_alternatives_temp"): # Nếu có từ lần submit lỗi trước ở bước PA
            form_data_alt_to_render = session.get("form_data_alternatives_temp")


    try:
        # Ưu tiên 1: Nếu có cờ import ma trận TC riêng và form_data_criteria_temp đã được cập nhật
        # (Logic này có thể không cần thiết nếu bạn đã gộp nút import TC riêng vào import_excel_route)
        if session.get("specific_criteria_matrix_imported") and session.get("form_data_criteria_temp"):
            print("DEBUG: calculate_criteria - Using criteria matrix from specific import (form_data_criteria_temp).")
            form_data_criteria_to_render = session.pop("form_data_criteria_temp") 
            session.pop("specific_criteria_matrix_imported", None) 

            criteria_matrix_parsed_np = np.ones((n_criteria, n_criteria), dtype=float)
            for i in range(n_criteria):
                for j in range(n_criteria):
                    value_str = form_data_criteria_to_render.get(f"matrix[{i}][{j}]", "1" if i == j else None)
                    if value_str is None and i < j:
                        raise ValueError(f"Lỗi dữ liệu (sau import TC): Thiếu ô TC [{criteria_names_ordered[i]}] vs [{criteria_names_ordered[j]}]")
                    value_str = value_str if value_str is not None else ("1" if i == j else "")
                    if i == j: criteria_matrix_parsed_np[i,j] = parse_saaty_value(value_str.strip(), True)
                    elif i < j:
                        pv = parse_saaty_value(value_str.strip()); criteria_matrix_parsed_np[i,j] = pv
                        criteria_matrix_parsed_np[j,i] = 1.0/pv if abs(pv)>1e-9 else np.inf
            session['imported_criteria_matrix'] = criteria_matrix_parsed_np.tolist() # Cập nhật lại session import đầy đủ
        
        # Ưu tiên 2: Nếu có import đầy đủ từ Excel (Bước 0 hoặc nút "Cập nhật")
        elif session.get('data_imported_from_excel') and 'imported_criteria_matrix' in session:
            print("DEBUG: calculate_criteria - Using criteria matrix from full Excel import.")
            criteria_matrix_parsed_np = np.array(session['imported_criteria_matrix'])
            form_data_criteria_to_render = convert_numpy_matrix_to_form_data(criteria_matrix_parsed_np, "matrix")
            # Lưu lại vào form_data_criteria_temp để nếu có lỗi ở bước sau còn hiển thị
            session["form_data_criteria_temp"] = form_data_criteria_to_render
        
        # Ưu tiên 3: Lấy từ form HTML (người dùng nhập tay hoặc sửa sau import)
        else:
            print("DEBUG: calculate_criteria - Parsing criteria matrix from HTML form.")
            form_data_from_html = request.form.to_dict()
            session["form_data_criteria_temp"] = form_data_from_html # Lưu lại form hiện tại
            form_data_criteria_to_render = form_data_from_html
            criteria_matrix_parsed_np = np.ones((n_criteria, n_criteria), dtype=float)
            for i in range(n_criteria):
                for j in range(n_criteria):
                    value_str = request.form.get(f"matrix[{i}][{j}]", "1" if i == j else None)
                    if value_str is None and i < j:
                        raise ValueError(f"Vui lòng điền ô Tiêu chí [{criteria_names_ordered[i]}] vs [{criteria_names_ordered[j]}]")
                    value_str = value_str if value_str is not None else ("1" if i == j else "")
                    if i == j: criteria_matrix_parsed_np[i,j] = parse_saaty_value(value_str.strip(), True)
                    elif i < j:
                        pv = parse_saaty_value(value_str.strip()); criteria_matrix_parsed_np[i,j] = pv
                        criteria_matrix_parsed_np[j,i] = 1.0/pv if abs(pv)>1e-9 else np.inf
        
        ahp_results_criteria = calculate_ahp(criteria_matrix_parsed_np, "Ma trận Tiêu chí")
        session["criteria_ahp_results"] = ahp_results_criteria
        sorted_crit_weights_template = get_sorted_criteria_with_weights(ahp_results_criteria, db_criteria_tuples)

        if ahp_results_criteria.get("error"):
            flash(ahp_results_criteria["error"], "error")
        else:
            is_consistent = ahp_results_criteria.get("is_consistent", False)
            cr_val = ahp_results_criteria.get("CR")
            cr_disp = f"{cr_val:.4f}" if isinstance(cr_val, (float,int)) else str(cr_val)
            if is_consistent:
                flash(f"Ma trận tiêu chí NHẤT QUÁN (CR = {cr_disp}). Đã lưu thông số.", "success")
                if ahp_results_criteria.get("weights"):
                    save_criteria_weights(current_session_db_id, db_criteria_tuples, ahp_results_criteria["weights"])
                # Chỉ lưu ma trận nếu không phải từ import hoặc nếu muốn ghi đè
                # Cân nhắc: nếu người dùng import rồi sửa trên form, có nên lưu lại ma trận đã sửa không?
                # Hiện tại: chỉ lưu nếu không phải từ 'data_imported_from_excel' (import toàn bộ)
                # và không phải từ 'specific_criteria_matrix_imported' (đã bị xóa ở trên)
                if not session.get('data_imported_from_excel'): 
                     save_criteria_comparison_matrix(current_session_db_id, db_criteria_tuples, criteria_matrix_parsed_np)
            else:
                flash(f"Ma trận tiêu chí KHÔNG nhất quán (CR = {cr_disp}). Thông số KHÔNG được lưu. Vui lòng sửa lại.", "error")
        print("AHP results:", ahp_results_criteria)
        # -- Phần render template đã được tích hợp logic chuẩn bị form_data_alt_to_render ở đầu --
        return render_template(
            "matrix.html",
            criteria=db_criteria_tuples,
            alternatives=db_alternatives_tuples,
            criteria_results=ahp_results_criteria,
            sorted_criteria_with_weights_for_template=sorted_crit_weights_template,
            show_input_forms=False,
            show_ahp_steps=True, 
            show_alternatives_form=(
                ahp_results_criteria.get("is_consistent", False)
                if ahp_results_criteria # Chỉ hiện form PA nếu TC nhất quán
                else False
            ),
            form_data_criteria=form_data_criteria_to_render,
            form_data_alternatives=form_data_alt_to_render, # <<<< Đã được chuẩn bị
            imported_data=session.get('data_imported_from_excel', False)
        )
    except ValueError as e: # Lỗi từ parse_saaty_value hoặc raise ValueError ở trên
        flash(f"Lỗi nhập liệu ma trận tiêu chí: {e}", "error")
        # Tạo kết quả lỗi giả để hiển thị
        ahp_err_res = {
            "error": str(e), "CR": "Lỗi nhập liệu", "is_consistent": False, 
            "weights": [1/n_criteria]*n_criteria if n_criteria > 0 else []
        }
        session["criteria_ahp_results"] = ahp_err_res # Lưu kết quả lỗi vào session
        
        # Khi render lỗi, vẫn truyền form_data_alternatives đã chuẩn bị
        return render_template(
            "matrix.html",
            criteria=db_criteria_tuples,
            alternatives=db_alternatives_tuples,
            criteria_results=ahp_err_res, # Kết quả lỗi
            # sorted_criteria_with_weights_for_template sẽ rỗng hoặc dựa trên weights lỗi
            sorted_criteria_with_weights_for_template=get_sorted_criteria_with_weights(ahp_err_res, db_criteria_tuples),
            show_input_forms=False,
            show_ahp_steps=True,
            show_alternatives_form=False, # Vì có lỗi ở ma trận TC
            form_data_criteria=form_data_criteria_to_render, # form_data_criteria_to_render đã được set ở đầu try
            form_data_alternatives=form_data_alt_to_render, # <<<< Truyền đi
            imported_data=session.get('data_imported_from_excel', False)
        )
    except Exception as e_gen:
        flash(f"Lỗi không xác định khi tính toán tiêu chí: {e_gen}", "error")
        traceback.print_exc()
        # Khi render lỗi chung, vẫn truyền form_data_alternatives đã chuẩn bị
        return render_template(
            "matrix.html",
            criteria=db_criteria_tuples,
            alternatives=db_alternatives_tuples,
            criteria_results=None, # Không có kết quả cụ thể
            sorted_criteria_with_weights_for_template=[],
            show_input_forms=False,
            show_ahp_steps=True,
            show_alternatives_form=False,
            form_data_criteria=form_data_criteria_to_render, # Có thể là rỗng nếu lỗi xảy ra sớm
            form_data_alternatives=form_data_alt_to_render, # <<<< Truyền đi
            imported_data=session.get('data_imported_from_excel', False)
        )
@app.route("/calculate_final", methods=["POST"])
def calculate_final_route():
    flask_session_id_value = session.get("flask_session_id")
    current_session_db_id = get_or_create_session_db_id(flask_session_id_value)
    db_criteria_tuples = session.get("current_criteria_tuples")
    db_alternatives_tuples = session.get("current_alternatives_tuples")
    criteria_ahp_results = session.get("criteria_ahp_results")

    if not all([current_session_db_id, db_criteria_tuples, db_alternatives_tuples, criteria_ahp_results]):
        flash("Lỗi phiên làm việc hoặc thiếu dữ liệu cần thiết. Vui lòng bắt đầu lại.", "error")
        return redirect(url_for("home"))

    form_data_criteria_render = session.get("form_data_criteria_temp", {})
    if session.get('data_imported_from_excel') and 'imported_criteria_matrix' in session and not form_data_criteria_render:
         form_data_criteria_render = convert_numpy_matrix_to_form_data(np.array(session['imported_criteria_matrix']), "matrix")
    
    sorted_crit_weights_template = get_sorted_criteria_with_weights(criteria_ahp_results, db_criteria_tuples)

    if not criteria_ahp_results.get("is_consistent", False) or not criteria_ahp_results.get("weights"):
        flash("Lỗi: Ma trận tiêu chí không nhất quán hoặc thiếu trọng số. Vui lòng sửa ở Bước 1.", "error")
        return render_template(
            "matrix.html", criteria=db_criteria_tuples, alternatives=db_alternatives_tuples,
            criteria_results=criteria_ahp_results, sorted_criteria_with_weights_for_template=sorted_crit_weights_template,
            show_input_forms=False, show_ahp_steps=True, show_alternatives_form=False,
            form_data_criteria=form_data_criteria_render, imported_data=session.get('data_imported_from_excel', False)
        )

    criteria_weights_vector = np.array(criteria_ahp_results["weights"])
    n_criteria = len(db_criteria_tuples)
    n_alternatives = len(db_alternatives_tuples)
    criteria_names_ordered = [name for _, name in db_criteria_tuples]
    alternative_names_ordered = [name for _, name in db_alternatives_tuples]
    
    alternative_local_scores_matrix_np = np.zeros((n_alternatives, n_criteria))
    any_alt_matrix_inconsistent = False
    any_critical_processing_error = False
    form_data_alternatives_render = {}
    alternative_ahp_results_by_crit_idx = {}

    try:
        if session.get('data_imported_from_excel') and 'imported_alternative_matrices' in session:
            imported_alt_matrices_dict_list = session['imported_alternative_matrices']
            temp_form_data_alts = {}
            for crit_idx_str, alt_matrix_list in imported_alt_matrices_dict_list.items():
                crit_idx = int(crit_idx_str)
                alt_matrix_parsed_np = np.array(alt_matrix_list)
                temp_form_data_alts.update(convert_numpy_matrix_to_form_data(alt_matrix_parsed_np, "alt_matrix", crit_idx))
                
                matrix_name_err = f"Phương án (TC: {criteria_names_ordered[crit_idx]})"
                alt_ahp_res = calculate_ahp(alt_matrix_parsed_np, matrix_name_err)
                alternative_ahp_results_by_crit_idx[str(crit_idx)] = alt_ahp_res
                if alt_ahp_res.get("error"): any_critical_processing_error = True
                elif not alt_ahp_res.get("is_consistent", False): any_alt_matrix_inconsistent = True
                if alt_ahp_res.get("weights"): alternative_local_scores_matrix_np[:, crit_idx] = np.array(alt_ahp_res["weights"])
                else: any_critical_processing_error = True; alternative_local_scores_matrix_np[:, crit_idx] = 0
            form_data_alternatives_render = temp_form_data_alts
        else:
            form_data_from_html = request.form.to_dict()
            session["form_data_alternatives_temp"] = form_data_from_html
            form_data_alternatives_render = form_data_from_html
            for crit_idx in range(n_criteria):
                alt_matrix_parsed_np = np.ones((n_alternatives, n_alternatives), dtype=float)
                input_err_flag_alt = False
                for i in range(n_alternatives):
                    for j in range(n_alternatives):
                        val_str = request.form.get(f"alt_matrix[{crit_idx}][{i}][{j}]", "1" if i==j else None)
                        if val_str is None and i < j: input_err_flag_alt = True; flash(f"Lỗi nhập liệu Phương án cho TC {criteria_names_ordered[crit_idx]}", "error"); break
                        val_str = val_str if val_str is not None else ("1" if i==j else "")
                        try:
                            if i==j: alt_matrix_parsed_np[i,j] = parse_saaty_value(val_str.strip(), True)
                            elif i < j: pv = parse_saaty_value(val_str.strip()); alt_matrix_parsed_np[i,j] = pv; alt_matrix_parsed_np[j,i] = 1.0/pv if abs(pv)>1e-9 else np.inf
                        except ValueError as e_parse_alt: input_err_flag_alt = True; flash(f"Lỗi giá trị Phương án cho TC {criteria_names_ordered[crit_idx]}: {e_parse_alt}", "error"); break
                    if input_err_flag_alt: break
                
                if input_err_flag_alt: any_critical_processing_error = True; alt_ahp_res = {"error": "Lỗi nhập liệu ma trận Phương án."}
                else: alt_ahp_res = calculate_ahp(alt_matrix_parsed_np, f"Phương án (TC: {criteria_names_ordered[crit_idx]})")
                
                alternative_ahp_results_by_crit_idx[str(crit_idx)] = alt_ahp_res
                if alt_ahp_res.get("error"): any_critical_processing_error = True
                elif not alt_ahp_res.get("is_consistent", False): any_alt_matrix_inconsistent = True
                if alt_ahp_res.get("weights"): alternative_local_scores_matrix_np[:, crit_idx] = np.array(alt_ahp_res["weights"])
                else: any_critical_processing_error = True; alternative_local_scores_matrix_np[:, crit_idx] = 0
        
        session["alternative_crs_temp"] = alternative_ahp_results_by_crit_idx

        if any_critical_processing_error:
            for crit_idx_str, alt_ahp_res in alternative_ahp_results_by_crit_idx.items():
                if not alt_ahp_res.get("is_consistent", False):
                    crit_idx = int(crit_idx_str)
                    crit_name = criteria_names_ordered[crit_idx] if crit_idx < len(criteria_names_ordered) else f"Tiêu chí {crit_idx+1}"
                    cr_val = alt_ahp_res.get("CR", "N/A")
                    lambda_max = alt_ahp_res.get("lambdaMax", "N/A")
                    ci = alt_ahp_res.get("ci", "N/A")
                    ri = alt_ahp_res.get("RI", "N/A")
                    flash(
                        f"Ma trận phương án cho <b>{crit_name}</b> KHÔNG nhất quán (λmax={lambda_max:.4f} | CI={ci:.4f} | RI={ri:.2f} | CR={cr_val:.4f}). Vui lòng nhập lại.",
                        "error"
                    )
        
        if any_alt_matrix_inconsistent:
            flash("ÍT NHẤT MỘT ma trận Phương án KHÔNG nhất quán. Kết quả có thể không đáng tin cậy.", "warning")
            return render_template(
            "matrix.html",
            criteria=db_criteria_tuples,
            alternatives=db_alternatives_tuples,
            criteria_results=criteria_ahp_results,
            sorted_criteria_with_weights_for_template=sorted_crit_weights_template,
            show_input_forms=False,
            show_ahp_steps=True,
            show_alternatives_form=True,
            form_data_criteria=form_data_criteria_render,
            form_data_alternatives=form_data_alternatives_render,
            alternative_crs=alternative_ahp_results_by_crit_idx,
            imported_data=session.get('data_imported_from_excel', False)
        )
        final_scores_np = np.dot(alternative_local_scores_matrix_np, criteria_weights_vector)
        total_final_score = np.sum(final_scores_np)
        if abs(total_final_score) > 1e-9 and not math.isclose(total_final_score, 1.0, abs_tol=1e-6):
            final_scores_np /= total_final_score
        
        save_alternative_scores(current_session_db_id, db_alternatives_tuples, db_criteria_tuples, alternative_local_scores_matrix_np)
        
        result_data = {
            "criteria_names": criteria_names_ordered, "criteria_weights": criteria_weights_vector.tolist(),
            "alternatives": alternative_names_ordered, "alternative_scores": final_scores_np.tolist(),
            "ranked_alternatives": sorted(list(zip(alternative_names_ordered, final_scores_np.tolist())), key=lambda x: x[1], reverse=True),
            "cr_criteria": criteria_ahp_results.get("CR"), "is_consistent_criteria": criteria_ahp_results.get("is_consistent"),
            "alternative_crs": alternative_ahp_results_by_crit_idx,
            "local_alternative_weights_matrix": alternative_local_scores_matrix_np.tolist(),
        }
        charts_data = generate_charts_to_files(result_data.get("ranked_alternatives"), result_data.get("criteria_names"), result_data.get("criteria_weights"))
        result_data["charts"] = charts_data
        saved_analysis_id = save_ahp_analysis(current_session_db_id, result_data)
        if saved_analysis_id:
            flash(f"Kết quả AHP đã được lưu (ID: {saved_analysis_id}).", "success")
            result_data["analysis_id_for_report"] = saved_analysis_id
        else: flash("Lưu kết quả AHP thất bại.", "error")

        keys_to_clean_final = ["criteria_ahp_results", "form_data_criteria_temp", "form_data_alternatives_temp", "alternative_crs_temp",
                               "data_imported_from_excel", "imported_criteria_matrix", "imported_alternative_matrices"]
        # Giữ lại current_criteria_tuples và current_alternatives_tuples nếu muốn bắt đầu lại với cùng TC/PA
        for key in keys_to_clean_final: session.pop(key, None)
        
        return render_template("result.html", **result_data)

    except Exception as e_final:
        flash(f"Lỗi máy chủ không xác định khi tính toán cuối cùng: {e_final}", "error"); traceback.print_exc()
        return render_template(
            "matrix.html", criteria=db_criteria_tuples, alternatives=db_alternatives_tuples,
            criteria_results=criteria_ahp_results, sorted_criteria_with_weights_for_template=sorted_crit_weights_template,
            show_input_forms=False, show_ahp_steps=True, show_alternatives_form=True,
            form_data_criteria=form_data_criteria_render, form_data_alternatives=form_data_alternatives_render,
            alternative_crs=session.get("alternative_crs_temp"), imported_data=session.get('data_imported_from_excel', False)
        )

# --- generate_charts_to_files, create_excel_report, download_excel_report, download_pdf_report, view_result_route, history_list_route (GIỮ NGUYÊN) ---
def generate_charts_to_files(
    ranked_alternatives, criteria_names, criteria_weights, unique_prefix=""
):
    chart_filenames = {}
    try:
        font_path_for_mpl = os.path.join(app.static_folder, "fonts", "ttf", "DejaVuSans.ttf")
        if os.path.exists(font_path_for_mpl):
             plt.rcParams['font.sans-serif'] = ['DejaVu Sans']
             plt.rcParams['font.family'] = "sans-serif"
        else:
            print(f"CẢNH BÁO [generate_charts]: Không tìm thấy font DejaVuSans.ttf cho Matplotlib.")
    except Exception as e_font:
        print(f"Lỗi khi thiết lập font cho Matplotlib: {e_font}")

    plt.style.use("seaborn-v0_8-whitegrid")
    base_filename = f"{unique_prefix}_{str(uuid.uuid4())[:8]}"

    if (criteria_names and criteria_weights and len(criteria_names) == len(criteria_weights)):
        try:
            sorted_data = sorted(zip(criteria_weights, criteria_names), key=lambda x: x[0], reverse=True)
            sorted_weights = [item[0] for item in sorted_data]
            sorted_names = [item[1] for item in sorted_data]
            fig1, ax1 = plt.subplots(figsize=(10, max(6, len(sorted_names) * 0.5)))
            palette = sns.color_palette("viridis", len(sorted_names))
            bars = ax1.barh(sorted_names, sorted_weights, color=palette, edgecolor="black")
            ax1.set_title("Trọng số của các Tiêu chí", fontsize=16, pad=20, fontweight="bold")
            ax1.set_xlabel("Trọng số", fontsize=12, fontweight="bold")
            ax1.invert_yaxis()
            for bar in bars:
                width = bar.get_width()
                ax1.text(width + 0.005, bar.get_y() + bar.get_height()/2.0, f"{width:.3f}", ha="left", va="center", fontsize=9)
            ax1.set_xlim(0, max(sorted_weights) * 1.15 if sorted_weights else 1)
            fig1.tight_layout(pad=2.0)
            filename_crit = f"{base_filename}_criteria_weights.png"
            filepath_crit = os.path.join(CHARTS_FOLDER, filename_crit)
            fig1.savefig(filepath_crit, format="png", dpi=150, bbox_inches="tight")
            plt.close(fig1)
            chart_filenames["criteria_weights_chart_file"] = filename_crit
        except Exception as e:
            print(f"Lỗi khi tạo biểu đồ trọng số tiêu chí: {e}"); traceback.print_exc()
            chart_filenames["criteria_weights_chart_file"] = None

    if ranked_alternatives:
        try:
            alt_names = [item[0] for item in ranked_alternatives]
            alt_scores = [item[1] for item in ranked_alternatives]
            valid_alt_names = [name for i, name in enumerate(alt_names) if alt_scores[i] > 1e-6]
            valid_alt_scores = [score for score in alt_scores if score > 1e-6]

            if not valid_alt_scores:
                chart_filenames["final_scores_chart_file"] = None
            else:
                fig2, ax2 = plt.subplots(figsize=(9, 9), subplot_kw=dict(aspect="equal"))
                num_colors = len(valid_alt_names)
                colors_palette = sns.color_palette("husl", num_colors) if num_colors > 0 else []
                wedges, texts, autotexts = ax2.pie(
                    valid_alt_scores, autopct="%1.1f%%", startangle=90, colors=colors_palette,
                    pctdistance=0.85, wedgeprops=dict(width=0.4, edgecolor="w"))
                for autotext in autotexts: autotext.set_color("black"); autotext.set_fontsize(9); autotext.set_fontweight("bold")
                ax2.set_title("Tỷ lệ Điểm số các Phương án", fontsize=16, pad=20, fontweight="bold")
                ax2.legend(wedges, valid_alt_names, title="Phương án", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1), fontsize=10, title_fontsize="12")
                fig2.tight_layout(rect=[0, 0, 0.8, 1])
                filename_alt = f"{base_filename}_alternatives_pie.png"
                filepath_alt = os.path.join(CHARTS_FOLDER, filename_alt)
                fig2.savefig(filepath_alt, format="png", dpi=150, bbox_inches="tight")
                plt.close(fig2)
                chart_filenames["final_scores_chart_file"] = filename_alt
        except Exception as e:
            print(f"Lỗi khi tạo biểu đồ tròn điểm số phương án: {e}"); traceback.print_exc()
            chart_filenames["final_scores_chart_file"] = None
    return chart_filenames

def create_excel_report(analysis_data):
    """Tạo file Excel từ dữ liệu phân tích AHP."""
    wb = Workbook()
    # Xóa sheet mặc định nếu có
    if "Sheet" in wb.sheetnames:
        std = wb.get_sheet_by_name("Sheet")
        wb.remove_sheet(std)

    header_font = Font(bold=True, name='Calibri', size=12)
    title_font = Font(bold=True, name='Calibri', size=14, underline='single')
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')

    # --- Sheet 1: Trọng số Tiêu chí ---
    ws_crit_weights = wb.create_sheet("TrongSoTieuChi")
    today_str = datetime.now().strftime("Ngày tạo báo cáo: %d/%m/%Y %H:%M")
    ws_crit_weights.cell(row=2, column=1, value=today_str).font = Font(italic=True, size=11)
    ws_crit_weights.cell(row=1, column=1, value="TRỌNG SỐ VÀ TỶ SỐ NHẤT QUÁN CỦA TIÊU CHÍ").font = title_font
    ws_crit_weights.merge_cells('A1:C1')
    ws_crit_weights['A1'].alignment = center_alignment
    
    row = 3
    headers_cw = ["STT", "Tiêu chí", "Trọng số"]
    for col_idx, header_text in enumerate(headers_cw, 1):
        cell = ws_crit_weights.cell(row=row, column=col_idx, value=header_text)
        cell.font = header_font
        cell.alignment = center_alignment
    row += 1

    if analysis_data.get('criteria_names') and analysis_data.get('criteria_weights'):
        for i, name in enumerate(analysis_data['criteria_names']):
            weight = analysis_data['criteria_weights'][i]
            ws_crit_weights.cell(row=row + i, column=1, value=i + 1)
            ws_crit_weights.cell(row=row + i, column=2, value=name).alignment = left_alignment
            ws_crit_weights.cell(row=row + i, column=3, value=round(weight, 5) if isinstance(weight, (int, float)) else weight).number_format = '0.00000'
        row += len(analysis_data['criteria_names'])
    row += 1
    ws_crit_weights.cell(row=row, column=1, value="CR Ma trận Tiêu chí:").font = header_font
    cr_crit_val = analysis_data.get('cr_criteria')
    ws_crit_weights.cell(row=row, column=2, value=round(cr_crit_val, 4) if isinstance(cr_crit_val, (int, float)) else cr_crit_val).number_format = '0.0000'
    is_cons_crit_val = analysis_data.get('is_consistent_criteria')
    ws_crit_weights.cell(row=row, column=3, value="(Nhất quán)" if is_cons_crit_val else "(Không nhất quán)" if is_cons_crit_val is False else "").font = Font(italic=True)

    # Sửa auto-size cột tránh lỗi MergedCell
    for col_cells in ws_crit_weights.columns:
        max_len = 0
        col_letter = None
        for cell in col_cells:
            if not isinstance(cell, MergedCell):
                col_letter = cell.column_letter
                break
        if not col_letter:
            continue
        for cell in col_cells:
            try:
                if len(str(cell.value)) > max_len: max_len = len(str(cell.value))
            except: pass
        ws_crit_weights.column_dimensions[col_letter].width = max_len + 3

    # --- Sheet 2: Chi tiết AHP của Phương án theo từng Tiêu chí ---
    if analysis_data.get('alternative_crs') and analysis_data.get('criteria_names') and analysis_data.get('alternatives'):
        ws_pa_details = wb.create_sheet("ChiTietTrongSoPA")
        current_row = 1
        alternatives_list = analysis_data.get('alternatives', [])
        criteria_list = analysis_data.get('criteria_names', [])

        ws_pa_details.cell(row=current_row, column=1, value="PHÂN TÍCH CHI TIẾT PHƯƠNG ÁN THEO TỪNG TIÊU CHÍ").font = title_font
        ws_pa_details.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max(4, len(alternatives_list) + 1))
        ws_pa_details.cell(row=current_row, column=1).alignment = center_alignment
        current_row += 2

        for crit_idx_str, alt_ahp_detail_dict in analysis_data.get('alternative_crs', {}).items():
            if isinstance(alt_ahp_detail_dict, dict): # Đảm bảo là dict
                crit_idx = int(crit_idx_str)
                crit_name = criteria_list[crit_idx] if crit_idx < len(criteria_list) else f"Tiêu chí Index {crit_idx}"
                
                ws_pa_details.cell(row=current_row, column=1, value=f"Tiêu chí: {crit_name}").font = Font(bold=True, size=13)
                ws_pa_details.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
                current_row += 1

                headers_pa = ["Phương án", "Trọng số cục bộ", "WSV", "CV"]
                for col_num, header_text in enumerate(headers_pa, 1):
                    ws_pa_details.cell(row=current_row, column=col_num, value=header_text).font = header_font
                    ws_pa_details.cell(row=current_row, column=col_num).alignment = center_alignment
                current_row += 1

                pa_weights = alt_ahp_detail_dict.get('weights', [])
                pa_wsv = alt_ahp_detail_dict.get('wsv', [])
                pa_cv = alt_ahp_detail_dict.get('cv', [])

                for i, alt_name_disp in enumerate(alternatives_list):
                    ws_pa_details.cell(row=current_row + i, column=1, value=alt_name_disp).alignment = left_alignment
                    
                    weight_val = pa_weights[i] if i < len(pa_weights) else 'N/A'
                    ws_pa_details.cell(row=current_row + i, column=2, value=round(weight_val,4) if isinstance(weight_val, (float, int)) else weight_val).number_format = '0.0000'
                    
                    wsv_val = pa_wsv[i] if i < len(pa_wsv) else 'N/A'
                    ws_pa_details.cell(row=current_row + i, column=3, value=round(wsv_val,4) if isinstance(wsv_val, (float, int)) else wsv_val).number_format = '0.0000'

                    cv_val = pa_cv[i] if i < len(pa_cv) else 'N/A'
                    ws_pa_details.cell(row=current_row + i, column=4, value=round(cv_val,4) if isinstance(cv_val, (float, int)) else cv_val).number_format = '0.0000'

                current_row += len(alternatives_list)
                
                cr_pa_val = alt_ahp_detail_dict.get('CR')
                is_cons_pa = alt_ahp_detail_dict.get('is_consistent')
                ws_pa_details.cell(row=current_row, column=1, value="CR:").font = Font(bold=True)
                ws_pa_details.cell(row=current_row, column=2, value=round(cr_pa_val, 4) if isinstance(cr_pa_val, (float, int)) else cr_pa_val).number_format = '0.0000'
                ws_pa_details.cell(row=current_row, column=3, value="(Nhất quán)" if is_cons_pa else "(Không nhất quán)" if is_cons_pa is False else "").font = Font(italic=True)
                current_row += 2
        
        for col_cells in ws_pa_details.columns:
            max_len = 0
            col_letter = None
            for cell in col_cells:
                if not isinstance(cell, MergedCell):
                    col_letter = cell.column_letter
                    break
            if not col_letter:
                continue
            for cell in col_cells:
                try:
                    if len(str(cell.value)) > max_len: max_len = len(str(cell.value))
                except: pass
            ws_pa_details.column_dimensions[col_letter].width = max_len + 3

    # --- Sheet 3: Bảng tổng hợp Trọng số cục bộ của PA theo tất cả Tiêu chí ---
    if analysis_data.get('local_alternative_weights_matrix') and analysis_data.get('alternatives') and analysis_data.get('criteria_names'):
        ws_local_summary = wb.create_sheet("TongHopTrongSoPATheoTC")
        ws_local_summary.cell(row=1, column=1, value="BẢNG TỔNG HỢP TRỌNG SỐ CỤC BỘ CỦA PHƯƠNG ÁN THEO TIÊU CHÍ").font = title_font
        ws_local_summary.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(analysis_data.get('criteria_names',[])) + 1)
        ws_local_summary.cell(row=1, column=1).alignment = center_alignment
        
        row_ls = 3
        ws_local_summary.cell(row=row_ls, column=1, value="Phương án \\ Tiêu chí").font = header_font
        for j_idx, crit_name_val in enumerate(analysis_data['criteria_names']):
            ws_local_summary.cell(row=row_ls, column=j_idx + 2, value=crit_name_val).font = header_font
            ws_local_summary.cell(row=row_ls, column=j_idx + 2).alignment = center_alignment
        row_ls += 1
        
        for i_idx, alt_name_val in enumerate(analysis_data['alternatives']):
            ws_local_summary.cell(row=row_ls + i_idx, column=1, value=alt_name_val).font = Font(bold=True)
            ws_local_summary.cell(row=row_ls + i_idx, column=1).alignment = left_alignment
            for j_idx, _ in enumerate(analysis_data['criteria_names']):
                score_val = analysis_data['local_alternative_weights_matrix'][i_idx][j_idx]
                ws_local_summary.cell(row=row_ls + i_idx, column=j_idx + 2, value=round(score_val, 5) if isinstance(score_val, (int, float)) else score_val).number_format = '0.00000'
        
        for col_cells in ws_local_summary.columns:
            max_len = 0
            col_letter = None
            for cell in col_cells:
                if not isinstance(cell, MergedCell):
                    col_letter = cell.column_letter
                    break
            if not col_letter:
                continue
            for cell in col_cells:
                try:
                    if len(str(cell.value)) > max_len: max_len = len(str(cell.value))
                except: pass
            ws_local_summary.column_dimensions[col_letter].width = max_len + 3

    # --- Sheet 4: Xếp hạng cuối cùng ---
    ws_ranking = wb.create_sheet("XepHangCuoiCung")
    ws_ranking.cell(row=1, column=1, value="XẾP HẠNG CÁC PHƯƠNG ÁN").font = title_font
    ws_ranking.merge_cells('A1:D1')
    ws_ranking['A1'].alignment = center_alignment

    row_r = 3
    headers_r = ["Hạng", "Phương án", "Điểm số Tổng hợp", "Phần trăm (%)"]
    for col_idx, header_text in enumerate(headers_r, 1):
        cell = ws_ranking.cell(row=row_r, column=col_idx, value=header_text)
        cell.font = header_font
        cell.alignment = center_alignment
    row_r += 1

    if analysis_data.get('ranked_alternatives'):
        for i, (alt_name, score) in enumerate(analysis_data['ranked_alternatives']):
            ws_ranking.cell(row=row_r + i, column=1, value=i + 1)
            ws_ranking.cell(row=row_r + i, column=2, value=alt_name).alignment = left_alignment
            ws_ranking.cell(row=row_r + i, column=3, value=round(score, 4) if isinstance(score, (int,float)) else score).number_format = '0.0000'
            ws_ranking.cell(row=row_r + i, column=4, value=round(score * 100, 2) if isinstance(score, (int,float)) else "N/A").number_format = '0.00"%"'
        row_r += len(analysis_data['ranked_alternatives'])
        row_r +=1
        ws_ranking.cell(row=row_r, column=1, value="Phương án tốt nhất:").font = header_font
        ws_ranking.cell(row=row_r, column=2, value=analysis_data['ranked_alternatives'][0][0]).font = Font(bold=True, color="008000")

    for col_cells in ws_ranking.columns:
        max_len = 0
        col_letter = None
        for cell in col_cells:
            if not isinstance(cell, MergedCell):
                col_letter = cell.column_letter
                break
        if not col_letter:
            continue
        for cell in col_cells:
            try:
                if len(str(cell.value)) > max_len: max_len = len(str(cell.value))
            except: pass
        ws_ranking.column_dimensions[col_letter].width = max_len + 3

    # Lưu vào bộ nhớ đệm
    excel_data_in_memory = io.BytesIO()
    wb.save(excel_data_in_memory)
    excel_data_in_memory.seek(0) # Đưa con trỏ về đầu stream
    return excel_data_in_memory.getvalue()

@app.route("/download_excel_report/<int:analysis_id>")
def download_excel_report(analysis_id):
    flask_session_id_value = session.get("flask_session_id")
    current_session_db_id = get_or_create_session_db_id(flask_session_id_value) if flask_session_id_value else None
    if not current_session_db_id:
        flash("Phiên làm việc không hợp lệ để tải báo cáo.", "error"); return redirect(url_for("history_list_route"))
    analysis_data = get_ahp_analysis_by_id(analysis_id, session_db_id_check=current_session_db_id)
    if not analysis_data:
        flash(f"Không tìm thấy dữ liệu phân tích ID {analysis_id} hoặc bạn không có quyền.", "error"); return redirect(url_for("history_list_route"))
    try:
        excel_bytes = create_excel_report(analysis_data)
        if excel_bytes is None: raise ValueError("Không thể tạo dữ liệu Excel.")
        return send_file(io.BytesIO(excel_bytes), as_attachment=True, download_name=f"AHP_BaoCao_PhanTich_{analysis_id}.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        print(f"Lỗi khi tạo/gửi file Excel: {e}"); traceback.print_exc()
        flash(f"Không thể tạo báo cáo Excel: {e}", "error"); return redirect(url_for("view_result_route", analysis_id=analysis_id))

@app.route("/download_pdf_report/<int:analysis_id>")
def download_pdf_report(analysis_id):
    print(f"DEBUG ReportLab: download_pdf_report called for analysis_id: {analysis_id}")
    flask_session_id_value = session.get("flask_session_id")
    current_session_db_id = get_or_create_session_db_id(flask_session_id_value) if flask_session_id_value else None
    if not current_session_db_id:
        flash("Phiên làm việc không hợp lệ để tải báo cáo.", "error"); return redirect(url_for("history_list_route"))
    analysis_data = get_ahp_analysis_by_id(analysis_id, session_db_id_check=current_session_db_id)
    if not analysis_data:
        flash(f"Không tìm thấy dữ liệu phân tích ID {analysis_id} hoặc bạn không có quyền.", "error"); return redirect(url_for("history_list_route"))

    chart_files_info = generate_charts_to_files(
        analysis_data.get("ranked_alternatives"), analysis_data.get("criteria_names"),
        analysis_data.get("criteria_weights"), unique_prefix=f"pdf_report_{analysis_id}")

    pdf_buffer = io.BytesIO()
    doc_width, doc_height = A4; margin = 0.75 * inch
    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4, rightMargin=margin, leftMargin=margin, topMargin=margin, bottomMargin=margin)
    story = []
    # styles = getSampleStyleSheet() # Không dùng trực tiếp nữa, định nghĩa hết
    
    # Định nghĩa styles
    style_report_title = ParagraphStyle('ReportTitle', fontName=FONT_NAME_REGULAR, fontSize=18, leading=22, textColor=colors.HexColor("#005555"), alignment=TA_CENTER, spaceAfter=12*pt)
    style_h2 = ParagraphStyle('CustomH2', fontName=FONT_NAME_REGULAR, fontSize=14, leading=18, textColor=colors.HexColor("#005555"), alignment=TA_LEFT, spaceBefore=15*pt, spaceAfter=8*pt)
    style_h3 = ParagraphStyle('CustomH3', fontName=FONT_NAME_REGULAR, fontSize=11, leading=14, textColor=colors.HexColor("#333333"), alignment=TA_LEFT, spaceBefore=12*pt, spaceAfter=6*pt)
    style_normal = ParagraphStyle('CustomNormal', fontName=FONT_NAME_REGULAR, fontSize=10, leading=14, textColor=colors.HexColor("#333333"), alignment=TA_LEFT, spaceAfter=4*pt)
    style_table_text = ParagraphStyle('TableText', parent=style_normal, fontSize=8.5, leading=11, alignment=TA_CENTER, fontName=FONT_NAME_REGULAR)
    style_table_header = ParagraphStyle('TableHeader', parent=style_table_text, fontName=FONT_NAME_REGULAR) # <b> sẽ làm đậm
    style_alt_name_cell = ParagraphStyle('AltNameCell', parent=style_table_text, fontName=FONT_NAME_REGULAR, alignment=TA_LEFT)
    style_highlight_header = ParagraphStyle('HighlightHeader', parent=style_table_header, fontName=FONT_NAME_REGULAR)
    style_highlight_score = ParagraphStyle('HighlightScore', parent=style_table_text, fontName=FONT_NAME_REGULAR)
    style_formula_note = ParagraphStyle('FormulaNote', fontName=FONT_NAME_REGULAR, fontSize=9, leading=12, textColor=colors.HexColor("#555555"), alignment=TA_LEFT, spaceBefore=6*pt, spaceAfter=10*pt)
    style_ahp_params_container = ParagraphStyle('AHPParamsContainer', fontName=FONT_NAME_REGULAR, fontSize=8.5, leading=11, textColor=colors.HexColor("#444444"), alignment=TA_LEFT, spaceBefore=4*pt)
    style_status_text = ParagraphStyle('StatusText', parent=style_normal, fontSize=9, fontName=FONT_NAME_REGULAR, spaceBefore=5*pt)

    story.append(Paragraph("<b>Tổng hợp kết quả phân tích AHP</b>", style_report_title))
    today_str = datetime.now().strftime("Ngày tạo báo cáo: %d/%m/%Y %H:%M")
    story.append(Paragraph(today_str, ParagraphStyle(
    name='ReportDate',
    parent=style_normal,
    fontSize=10,
    textColor=colors.HexColor("#888888"),
    alignment=TA_RIGHT,
    spaceAfter=8*pt,
)))
    has_charts = chart_files_info.get("criteria_weights_chart_file") or chart_files_info.get("final_scores_chart_file")
    if has_charts:
        story.append(Paragraph("<b>Trực quan hóa kết quả</b>", style_h2))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#cccccc"), spaceBefore=2, spaceAfter=8*pt))
        if chart_files_info.get("criteria_weights_chart_file"):
            story.append(Paragraph("<b>Biểu đồ trọng số các tiêu chí</b>", style_h3))
            img_path_crit = os.path.join(CHARTS_FOLDER, chart_files_info["criteria_weights_chart_file"])
            if os.path.exists(img_path_crit):
                try: img_crit = Image(img_path_crit, width=doc_width*0.75, height=doc_height*0.3); img_crit.hAlign='CENTER'; story.append(img_crit); story.append(Spacer(1,0.2*inch))
                except Exception as e_img_c: story.append(Paragraph(f"[Lỗi ảnh TC: {e_img_c}]", style_normal))
            else: story.append(Paragraph("[Không tìm thấy ảnh TC]", style_normal))
        if chart_files_info.get("final_scores_chart_file"):
            story.append(Paragraph("<b>Biểu đồ tỷ lệ các phương án</b>", style_h3))
            img_path_alt = os.path.join(CHARTS_FOLDER, chart_files_info["final_scores_chart_file"])
            if os.path.exists(img_path_alt):
                try: img_alt = Image(img_path_alt, width=doc_width*0.7, height=doc_height*0.3); img_alt.hAlign='CENTER'; story.append(img_alt); story.append(Spacer(1,0.2*inch))
                except Exception as e_img_a: story.append(Paragraph(f"[Lỗi ảnh PA: {e_img_a}]", style_normal))
            else: story.append(Paragraph("[Không tìm thấy ảnh PA]", style_normal))
        story.append(Spacer(1, 0.1*inch))

    if (analysis_data.get("alternative_crs") and analysis_data.get("criteria_names") and analysis_data.get("alternatives")):
        story.append(Paragraph("<b>Phân tích chi tiết phương án theo từng tiêu chí</b>", style_h2))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#cccccc"), spaceBefore=2, spaceAfter=8*pt))
        crits_list_pdf = analysis_data["criteria_names"]; alts_list_pdf = analysis_data["alternatives"]; alt_crs_dict_pdf = analysis_data["alternative_crs"]
        num_crits_detail_pdf = len(alt_crs_dict_pdf); curr_crit_detail_idx_pdf = 0
        for crit_idx_str, alt_ahp_dict in alt_crs_dict_pdf.items():
            curr_crit_detail_idx_pdf += 1; crit_idx = int(crit_idx_str)
            curr_crit_name_pdf = crits_list_pdf[crit_idx] if crit_idx < len(crits_list_pdf) else f"Tiêu chí {crit_idx+1}"
            crit_details_story_seg = [Paragraph(f"<b>Chi tiết cho Tiêu chí: {curr_crit_name_pdf}</b>", style_h3)]
            if (alt_ahp_dict and alt_ahp_dict.get("error") is None and all(k in alt_ahp_dict for k in ["weights","wsv","cv"])):
                tbl_data_alt_det = [[Paragraph(f"<b>{h}</b>",style_table_header) for h in ["Phương án","Trọng số","WSV","CV"]]]
                for i, alt_n in enumerate(alts_list_pdf):
                    w=alt_ahp_dict["weights"][i] if i < len(alt_ahp_dict["weights"]) else "N/A"
                    wsv=alt_ahp_dict["wsv"][i] if i < len(alt_ahp_dict["wsv"]) else "N/A"
                    cv=alt_ahp_dict["cv"][i] if i < len(alt_ahp_dict["cv"]) else "N/A"
                    row_det = [Paragraph(alt_n, style_alt_name_cell),
                               Paragraph(f"{w:.4f}" if isinstance(w,(float,int)) else str(w), style_table_text),
                               Paragraph(f"{wsv:.4f}" if isinstance(wsv,(float,int)) else str(wsv), style_table_text),
                               Paragraph(f"{cv:.4f}" if isinstance(cv,(float,int)) else str(cv), style_table_text)]
                    tbl_data_alt_det.append(row_det)
                avail_tbl_w = doc_width - margin*2 - (0.2*inch)
                col_w_alt_det = [avail_tbl_w*0.35, avail_tbl_w*0.23, avail_tbl_w*0.21, avail_tbl_w*0.21]
                tbl_alt_det_obj = Table(tbl_data_alt_det, colWidths=col_w_alt_det, repeatRows=1)
                tbl_alt_det_obj.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor("#e9ecef")),('GRID',(0,0),(-1,-1),1,colors.HexColor("#777777")),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('LEFTPADDING',(0,0),(-1,-1),3*pt),('RIGHTPADDING',(0,0),(-1,-1),3*pt),('TOPPADDING',(0,0),(-1,-1),2*pt),('BOTTOMPADDING',(0,0),(-1,-1),2*pt)]))
                crit_details_story_seg.append(tbl_alt_det_obj); crit_details_story_seg.append(Spacer(1,5*pt))
                lm=alt_ahp_dict.get("lambdaMax"); ci_v=alt_ahp_dict.get("ci"); ri_v=alt_ahp_dict.get("RI"); n_v=alt_ahp_dict.get("n"); cr_v=alt_ahp_dict.get("CR")
                lm_s=f"{lm:.4f}" if isinstance(lm,(float,int)) else "N/A"; ci_s=f"{ci_v:.4f}" if isinstance(ci_v,(float,int)) else "N/A"
                ri_s=f"{ri_v:.2f}" if isinstance(ri_v,(float,int)) else "N/A"; n_s=f"(n={n_v})" if isinstance(n_v,int) else ""
                cr_s=f"{cr_v:.4f}" if isinstance(cr_v,(float,int)) else str(cr_v if cr_v is not None else "N/A")
                st_txt=""; clr_hex=colors.black.hexval()
                if isinstance(cr_v,(float,int)): st_txt="(Nhất quán)" if alt_ahp_dict.get("is_consistent") else "(Không nhất quán)"; clr_hex=colors.green.hexval() if alt_ahp_dict.get("is_consistent") else colors.red.hexval()
                elif cr_v=="Lỗi nhập liệu" or alt_ahp_dict.get("error"): st_txt=f"({alt_ahp_dict.get('CR') or alt_ahp_dict.get('error')})"; clr_hex=colors.red.hexval()
                params_html = f"<b>Lambda_max:</b> {lm_s} | <b>CI:</b> {ci_s} | <b>RI:</b> {ri_s} {n_s}<br/><b>CR: {cr_s}</b> <font color='{clr_hex}'>{st_txt}</font>"
                crit_details_story_seg.append(Paragraph(params_html, style_ahp_params_container))
            elif alt_ahp_dict and alt_ahp_dict.get("error"):
                err_msg = f"Lỗi: {alt_ahp_dict['error']}" + (f" (CR: {alt_ahp_dict['CR']})" if alt_ahp_dict.get('CR') else "")
                crit_details_story_seg.append(Paragraph(err_msg, ParagraphStyle(parent=style_normal, textColor=colors.red)))
            else: crit_details_story_seg.append(Paragraph("Không có đủ dữ liệu chi tiết.", ParagraphStyle(parent=style_normal,textColor=colors.orange)))
            if curr_crit_detail_idx_pdf < num_crits_detail_pdf: crit_details_story_seg.extend([Spacer(1,0.1*inch),HRFlowable(width="80%",thickness=0.5,color=colors.HexColor("#aaaaaa"),dash=(2,2),hAlign='CENTER'),Spacer(1,0.1*inch)])
            else: crit_details_story_seg.append(Spacer(1,0.2*inch))
            story.append(KeepInFrame(doc_width-margin*2, 6*inch, crit_details_story_seg)) # Bỏ [0]
            
    story.append(PageBreak())
    story.append(Paragraph("<b>Tóm tắt và tổng hợp kết quả</b>", style_h2))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#cccccc"), spaceBefore=2, spaceAfter=8*pt))
    formula = "<b>Công thức:</b> Điểm Phương án = Sum ( [Trọng số Phương án theo tiêu chí] * [Trọng số tiêu chí] )"
    story.append(Paragraph(formula, style_formula_note))
    if (analysis_data.get("local_alternative_weights_matrix") and analysis_data.get("alternatives") and analysis_data.get("criteria_names")):
        story.append(Paragraph("<b>Bảng 1: Trọng số phương án theo tiêu chí</b>", style_h3))
        cn_b1=analysis_data["criteria_names"]; alts_b1=analysis_data["alternatives"]; lw_b1=analysis_data["local_alternative_weights_matrix"]
        hr_b1=[Paragraph("<b>Phương án\\TC</b>",style_highlight_header)]+[Paragraph(f"<b>{n}</b>",style_highlight_header) for n in cn_b1]
        td_b1=[hr_b1]
        for i, alt_n in enumerate(alts_b1):
            row_b1 = [Paragraph(alt_n, style_alt_name_cell)]
            for j in range(len(cn_b1)): v=lw_b1[i][j]; v_s=f"{v:.5f}" if isinstance(v,(float,int)) else "N/A"; row_b1.append(Paragraph(v_s,style_table_text))
            td_b1.append(row_b1)
        n_c_b1=len(cn_b1); fcw_b1=1.5*inch; rem_w_b1=doc_width-(margin*2)-fcw_b1; ocw_b1=(rem_w_b1/n_c_b1) if n_c_b1>0 else 0
        cw_b1=[fcw_b1]+[ocw_b1]*n_c_b1 if n_c_b1 > 0 else [doc_width-(margin*2)]
        t_b1=Table(td_b1,colWidths=cw_b1,repeatRows=1); t_b1.setStyle(TableStyle([('GRID',(0,0),(-1,-1),1,colors.HexColor("#777777")),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BACKGROUND',(0,0),(-1,0),colors.HexColor("#e9ecef"))]))
        story.append(t_b1); story.append(Spacer(1,0.2*inch))

    if analysis_data.get("criteria_weights") and analysis_data.get("criteria_names"):
        story.append(Paragraph("<b>Bảng 2: Trọng số các Tiêu chí</b>", style_h3))
        td_b2=[[Paragraph("<b>Tiêu chí</b>",style_highlight_header),Paragraph("<b>Trọng số</b>",style_highlight_header)]]
        for i, cn_b2 in enumerate(analysis_data["criteria_names"]):
            w_b2=analysis_data["criteria_weights"][i]; w_s_b2=f"{w_b2:.5f}" if isinstance(w_b2,(float,int)) else "N/A"
            td_b2.append([Paragraph(cn_b2,style_alt_name_cell),Paragraph(w_s_b2,style_table_text)])
        tw_b2=doc_width-(margin*2); cw_b2=[tw_b2*0.6,tw_b2*0.4]
        t_b2=Table(td_b2,colWidths=cw_b2,repeatRows=1); t_b2.setStyle(TableStyle([('GRID',(0,0),(-1,-1),1,colors.HexColor("#777777")),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BACKGROUND',(0,0),(-1,0),colors.HexColor("#e9ecef"))]))
        story.append(t_b2)
        cr_c_v=analysis_data.get("cr_criteria"); cr_c_s=f"{cr_c_v:.4f}" if isinstance(cr_c_v,(float,int)) else str(cr_c_v if cr_c_v is not None else "N/A")
        is_c_c=analysis_data.get("is_consistent_criteria"); con_t_c=""; con_c_c_hex=colors.black.hexval()
        if isinstance(cr_c_v,(float,int)): con_t_c="(Nhất quán)" if is_c_c else "(Không nhất quán)"; con_c_c_hex=(colors.green if is_c_c else colors.red).hexval()
        p_cr_c_txt=f"CR Ma trận Tiêu chí: <b>{cr_c_s}</b> <font color='{con_c_c_hex}'>{con_t_c}</font>"
        story.append(Paragraph(p_cr_c_txt,style_status_text)); story.append(Spacer(1,0.2*inch))

    if analysis_data.get("alternative_scores") and analysis_data.get("alternatives"):
        story.append(Paragraph("<b>Bảng 3: Điểm số Tổng hợp và Tỷ lệ Phương án</b>", style_h3))
        td_b3=[[Paragraph(f"<b>{h}</b>",style_highlight_header) for h in ["Phương án","Điểm Tổng hợp","Tỷ lệ (%)"]]]
        for i, alt_n_b3 in enumerate(analysis_data["alternatives"]):
            s_b3=analysis_data["alternative_scores"][i]; s_s_b3=f"{s_b3:.4f}" if isinstance(s_b3,(float,int)) else "N/A"
            p_s_b3=f"{(s_b3*100):.2f}%" if isinstance(s_b3,(float,int)) else "N/A"
            td_b3.append([Paragraph(alt_n_b3,style_alt_name_cell),Paragraph(s_s_b3,style_highlight_score),Paragraph(p_s_b3,style_highlight_score)])
        tw_b3=doc_width-(margin*2); cw_b3=[tw_b3*w for w in [0.4,0.3,0.3]]
        t_b3=Table(td_b3,colWidths=cw_b3,repeatRows=1); t_b3.setStyle(TableStyle([('GRID',(0,0),(-1,-1),1,colors.HexColor("#777777")),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BACKGROUND',(0,0),(-1,0),colors.HexColor("#e9ecef"))]))
        story.append(t_b3); story.append(Spacer(1,0.2*inch))

    if analysis_data.get("ranked_alternatives"):
        story.append(Paragraph("<b>Xếp hạng các Phương án</b>", style_h2))
        story.append(HRFlowable(width="100%",thickness=1,color=colors.HexColor("#cccccc"),spaceBefore=2,spaceAfter=8*pt))
        td_r=[[Paragraph(f"<b>{h}</b>",style_highlight_header) for h in ["Hạng","Phương án","Điểm số","Tỷ lệ (%)"]]]
        for i, (alt_n_r,s_r) in enumerate(analysis_data["ranked_alternatives"]):
            r_s=str(i+1); s_s_r=f"{s_r:.4f}" if isinstance(s_r,(float,int)) else "N/A"; p_s_r=f"{(s_r*100):.2f}%" if isinstance(s_r,(float,int)) else "N/A"
            td_r.append([Paragraph(r_s,style_table_text),Paragraph(alt_n_r,style_alt_name_cell),Paragraph(s_s_r,style_highlight_score),Paragraph(p_s_r,style_highlight_score)])
        tw_r=doc_width-(margin*2); cw_r=[0.7*inch, tw_r*0.45-0.7*inch, tw_r*0.25, tw_r*0.30]
        t_r=Table(td_r,colWidths=cw_r,repeatRows=1); t_r.setStyle(TableStyle([('GRID',(0,0),(-1,-1),1,colors.HexColor("#777777")),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('ALIGN',(0,0),(0,-1),'CENTER'),('BACKGROUND',(0,0),(-1,0),colors.HexColor("#e9ecef"))]))
        story.append(t_r)
        if analysis_data["ranked_alternatives"] and len(analysis_data["ranked_alternatives"])>0:
            b_alt_n,b_alt_s = analysis_data["ranked_alternatives"][0]
            b_alt_s_s = f"{b_alt_s:.4f}" if isinstance(b_alt_s,(float,int)) else "N/A"
            b_alt_txt=f"<b>Phương án đề xuất:</b> <font name='{FONT_NAME_BOLD}' color='{colors.HexColor('#28a745').hexval()}'>{b_alt_n}</font> (Điểm: <b>{b_alt_s_s}</b>)"
            story.append(Spacer(1,0.1*inch)); story.append(Paragraph(b_alt_txt,ParagraphStyle(name='BestAlt',parent=style_normal,fontSize=11,alignment=TA_CENTER,spaceBefore=10*pt,fontName=FONT_NAME_REGULAR)))
    try:
        doc.build(story)
    except Exception as e_build:
        print(f"ERROR ReportLab: Failed to build PDF: {e_build}"); traceback.print_exc()
        flash(f"Lỗi khi xây dựng PDF: {e_build}", "error")
        # Dọn dẹp file ảnh nếu lỗi build
        for f_val in chart_files_info.values():
            if f_val: 
                try: fp_del=os.path.join(CHARTS_FOLDER, f_val); os.remove(fp_del) if os.path.exists(fp_del) else None; 
                except Exception: pass
        return redirect(url_for("view_result_route", analysis_id=analysis_id))

    pdf_bytes = pdf_buffer.getvalue(); pdf_buffer.close()
    for f_val in chart_files_info.values(): # Dọn dẹp ảnh sau khi build thành công
        if f_val: 
            try: fp_del=os.path.join(CHARTS_FOLDER, f_val); os.remove(fp_del) if os.path.exists(fp_del) else None; 
            except Exception: pass
    
    if pdf_bytes: return Response(pdf_bytes, mimetype="application/pdf", headers={"Content-Disposition":f"attachment;filename=AHP_Report_Analysis_{analysis_id}.pdf"})
    else: flash("Không thể tạo PDF, dữ liệu rỗng.", "error"); return redirect(url_for("view_result_route", analysis_id=analysis_id))


@app.route("/history")
def history_list_route():
    flask_session_id_value = session.get("flask_session_id")
    session_db_id = get_or_create_session_db_id(flask_session_id_value) if flask_session_id_value else None
    if not session_db_id and not flask_session_id_value:
        flash("Vui lòng bắt đầu một phiên làm việc để xem lịch sử.", "info"); return redirect(url_for("home"))
    analyses_history = get_ahp_analyses_by_session_db_id(session_db_id) if session_db_id else []
    return render_template("history_list.html", analyses=analyses_history)


@app.route("/result_history/<int:analysis_id>")
def view_result_route(analysis_id):
    flask_session_id_value = session.get("flask_session_id")
    current_session_db_id = get_or_create_session_db_id(flask_session_id_value) if flask_session_id_value else None
    if not current_session_db_id:
        flash("Không thể xác thực phiên làm việc.", "error"); return redirect(url_for("history_list_route"))
    saved_analysis_data = get_ahp_analysis_by_id(analysis_id, session_db_id_check=current_session_db_id)
    if saved_analysis_data:
        charts_data = generate_charts_to_files(
            saved_analysis_data.get("ranked_alternatives"), saved_analysis_data.get("criteria_names"),
            saved_analysis_data.get("criteria_weights"), unique_prefix=f"history_{analysis_id}")
        saved_analysis_data["charts"] = charts_data
        saved_analysis_data["analysis_id_for_report"] = analysis_id
        return render_template("result.html", **saved_analysis_data)
    else:
        flash(f"Không tìm thấy kết quả phân tích ID {analysis_id} hoặc bạn không có quyền xem.", "error")
        return redirect(url_for("history_list_route"))


if __name__ == "__main__":
    app.run(debug=True)